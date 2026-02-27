"""
Modulo para leer archivos de comparativos desde:
1. Adjuntos de Gmail (archivos .xlsx)
2. Google Drive (carpetas/archivos enlazados en el correo)
3. Google Sheets (hojas de calculo en Drive)

Extrae:
- Monto CC (monto del ganador del comparativo)
- PPTO META HG (presupuesto meta, en pestaña "VS")

Estructura esperada de la pestaña "VS":
  - "PPTO META HG" es un HEADER DE SECCION que abarca varias columnas
  - Debajo del header hay sub-columnas: SUMINISTRO/SERVICIO, OBRA, V.U., SUB TOTAL
  - Al final de la hoja hay filas de totales: COSTO DIRECTO, SUB TOTAL, IGV, TOTAL (CON IGV)
  - El valor de PPTO META HG es el TOTAL (CON IGV) de esa seccion
"""
import os
import re
import io
import base64
import tempfile
from openpyxl import load_workbook

from config import TEMP_DIR


def extraer_datos_comparativo(gmail_service, drive_service, sheets_service, mensaje_id, cuerpo_fallback="", asunto="", thread_id=""):
    """
    Extrae Monto CC y PPTO META HG de un comparativo.
    1. Obtiene el mensaje completo (para adjuntos Y cuerpo completo)
    2. Intenta leer adjuntos Excel del mensaje
    3. Busca links de Drive en TODOS los mensajes del hilo (thread)
       - Las personas clave suelen responder "Se subio comparativo a la carpeta" con link de Drive
       - Si el link es una carpeta, busca el archivo Excel que coincida con el asunto

    Args:
        asunto: Asunto del correo, usado para matchear archivos en carpetas de Drive
        thread_id: ID del thread para buscar links en todas las respuestas
    """
    resultado = {
        "monto_cc": "No especificado",
        "ppto_meta_hg": "No especificado",
        "expediente": "No especificado",
    }

    # Obtener mensaje completo UNA sola vez
    try:
        msg = gmail_service.users().messages().get(
            userId="me", id=mensaje_id, format="full"
        ).execute()
    except Exception as e:
        print(f"    [WARN] Error obteniendo mensaje: {e}")
        return resultado

    payload = msg.get("payload", {})

    # 1. Intentar desde adjuntos Excel del mensaje original
    adjuntos = _buscar_adjuntos_recursivo(payload)
    for adj in adjuntos:
        filename = adj["filename"]
        if not any(filename.lower().endswith(ext) for ext in [".xlsx", ".xls", ".xlsm"]):
            continue

        try:
            attachment = gmail_service.users().messages().attachments().get(
                userId="me", messageId=mensaje_id, id=adj["attachmentId"]
            ).execute()
            file_data = base64.urlsafe_b64decode(attachment["data"])
            datos = _procesar_excel(file_data, filename)
            if datos:
                if datos.get("monto_cc") != "No especificado":
                    resultado["monto_cc"] = datos["monto_cc"]
                if datos.get("ppto_meta_hg") != "No especificado":
                    resultado["ppto_meta_hg"] = datos["ppto_meta_hg"]
                if datos.get("expediente") != "No especificado":
                    resultado["expediente"] = datos["expediente"]
            if resultado["monto_cc"] != "No especificado" and resultado["ppto_meta_hg"] != "No especificado":
                return resultado
        except Exception as e:
            print(f"    [WARN] Error con adjunto '{filename}': {e}")

    # 2. Buscar links de Drive en TODOS los mensajes del hilo
    #    Las personas clave responden con "Se subio comparativo a la carpeta" + link de Drive
    texto_completo_thread = ""

    if thread_id:
        try:
            thread = gmail_service.users().threads().get(
                userId="me", id=thread_id, format="full"
            ).execute()
            for thread_msg in thread.get("messages", []):
                thread_payload = thread_msg.get("payload", {})
                # Extraer texto y HTML de cada mensaje del hilo
                txt = _extraer_texto_de_payload(thread_payload)
                html = _extraer_html_de_payload(thread_payload)
                texto_completo_thread += (txt or "") + " " + (html or "") + " "

                # Tambien buscar adjuntos Excel en otros mensajes del hilo
                if resultado["ppto_meta_hg"] == "No especificado":
                    thread_adjuntos = _buscar_adjuntos_recursivo(thread_payload)
                    for adj in thread_adjuntos:
                        fname = adj["filename"]
                        if not any(fname.lower().endswith(ext) for ext in [".xlsx", ".xls", ".xlsm"]):
                            continue
                        try:
                            att = gmail_service.users().messages().attachments().get(
                                userId="me", messageId=thread_msg["id"], id=adj["attachmentId"]
                            ).execute()
                            fdata = base64.urlsafe_b64decode(att["data"])
                            datos = _procesar_excel(fdata, fname)
                            if datos and datos.get("ppto_meta_hg") != "No especificado":
                                resultado["ppto_meta_hg"] = datos["ppto_meta_hg"]
                                if resultado["monto_cc"] == "No especificado" and datos.get("monto_cc") != "No especificado":
                                    resultado["monto_cc"] = datos["monto_cc"]
                                if resultado["expediente"] == "No especificado" and datos.get("expediente") != "No especificado":
                                    resultado["expediente"] = datos["expediente"]
                                break
                        except Exception:
                            pass
        except Exception as e:
            print(f"    [WARN] Error leyendo thread: {e}")

    # Fallback: usar cuerpo del mensaje original si no hay thread
    if not texto_completo_thread:
        cuerpo_texto = _extraer_texto_de_payload(payload)
        cuerpo_html = _extraer_html_de_payload(payload)
        texto_completo_thread = (cuerpo_texto or "") + " " + (cuerpo_html or "") + " " + (cuerpo_fallback or "")

    # Buscar links de Drive en todo el texto recopilado
    drive_links = _extraer_drive_links(texto_completo_thread)
    for link in drive_links:
        try:
            datos_drive = _leer_desde_drive(drive_service, sheets_service, link, asunto=asunto)
            if datos_drive:
                if resultado["monto_cc"] == "No especificado" and datos_drive.get("monto_cc") != "No especificado":
                    resultado["monto_cc"] = datos_drive["monto_cc"]
                if resultado["ppto_meta_hg"] == "No especificado" and datos_drive.get("ppto_meta_hg") != "No especificado":
                    resultado["ppto_meta_hg"] = datos_drive["ppto_meta_hg"]
                if resultado["expediente"] == "No especificado" and datos_drive.get("expediente") != "No especificado":
                    resultado["expediente"] = datos_drive["expediente"]
            if resultado["monto_cc"] != "No especificado" and resultado["ppto_meta_hg"] != "No especificado":
                break
        except Exception as e:
            print(f"    [WARN] Error con Drive link: {e}")

    return resultado


# ============================================================================
# BUSQUEDA DE ADJUNTOS (recursivo para multipart anidado)
# ============================================================================

def _buscar_adjuntos_recursivo(payload, adjuntos=None):
    """Busca recursivamente todos los adjuntos en un payload de Gmail."""
    if adjuntos is None:
        adjuntos = []

    filename = payload.get("filename", "")
    attachment_id = payload.get("body", {}).get("attachmentId")

    if filename and attachment_id:
        adjuntos.append({
            "filename": filename,
            "attachmentId": attachment_id,
            "mimeType": payload.get("mimeType", ""),
        })

    for part in payload.get("parts", []):
        _buscar_adjuntos_recursivo(part, adjuntos)

    return adjuntos


# ============================================================================
# EXTRACCION DE CUERPO COMPLETO (texto y HTML)
# ============================================================================

def _extraer_texto_de_payload(payload):
    """Extrae texto plano completo del payload (recursivo)."""
    if payload.get("mimeType") == "text/plain" and payload.get("body", {}).get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="replace")

    for part in payload.get("parts", []):
        resultado = _extraer_texto_de_payload(part)
        if resultado:
            return resultado

    return ""


def _extraer_html_de_payload(payload):
    """Extrae HTML completo del payload (recursivo)."""
    if payload.get("mimeType") == "text/html" and payload.get("body", {}).get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode("utf-8", errors="replace")

    for part in payload.get("parts", []):
        resultado = _extraer_html_de_payload(part)
        if resultado:
            return resultado

    return ""


# ============================================================================
# PROCESAMIENTO DE EXCEL (openpyxl)
# ============================================================================

def _procesar_excel(file_data, filename="archivo.xlsx"):
    """
    Procesa un archivo Excel de comparativo.
    Prioriza la pestaña "VS" para PPTO META HG, EXPEDIENTE y Monto CC.
    """
    resultado = {"monto_cc": "No especificado", "ppto_meta_hg": "No especificado", "expediente": "No especificado"}

    try:
        wb = load_workbook(io.BytesIO(file_data), data_only=True)
        sheet_names_lower = [s.lower() for s in wb.sheetnames]

        # Buscar pestaña "VS" (prioridad exacta)
        vs_idx = None
        for i, name in enumerate(sheet_names_lower):
            if name.strip() == "vs":
                vs_idx = i
                break
        # Si no hay exacta, buscar que contenga "vs"
        if vs_idx is None:
            for i, name in enumerate(sheet_names_lower):
                if "vs" in name:
                    vs_idx = i
                    break

        if vs_idx is not None:
            ws = wb.worksheets[vs_idx]
            datos_vs = _leer_hoja_vs(ws)
            if datos_vs.get("ppto_meta_hg") != "No especificado":
                resultado["ppto_meta_hg"] = datos_vs["ppto_meta_hg"]
            if datos_vs.get("monto_cc") != "No especificado":
                resultado["monto_cc"] = datos_vs["monto_cc"]
            if datos_vs.get("expediente") != "No especificado":
                resultado["expediente"] = datos_vs["expediente"]

        # Si no encontramos en "VS", buscar en hojas con nombre relevante
        if resultado["ppto_meta_hg"] == "No especificado":
            for i, name in enumerate(sheet_names_lower):
                if any(kw in name for kw in ["comparativo", "resumen", "cuadro"]):
                    ws = wb.worksheets[i]
                    datos_hoja = _leer_hoja_vs(ws)
                    if datos_hoja.get("ppto_meta_hg") != "No especificado":
                        resultado["ppto_meta_hg"] = datos_hoja["ppto_meta_hg"]
                    if resultado["monto_cc"] == "No especificado" and datos_hoja.get("monto_cc") != "No especificado":
                        resultado["monto_cc"] = datos_hoja["monto_cc"]
                    if resultado["expediente"] == "No especificado" and datos_hoja.get("expediente") != "No especificado":
                        resultado["expediente"] = datos_hoja["expediente"]
                    if resultado["ppto_meta_hg"] != "No especificado":
                        break

        wb.close()

    except Exception as e:
        print(f"    [WARN] Error procesando Excel '{filename}': {e}")

    return resultado


def _leer_hoja_vs(ws):
    """
    Lee la hoja VS de un comparativo con formato de secciones.

    Estructura esperada:
    - Fila ~6: Headers de seccion: EXPEDIENTE | PROVEEDOR1 | PROVEEDOR2 | PPTO META HG
    - Fila ~7: Sub-headers: V.U. | SUB TOTAL | OBS | ... | V.U. | SUB TOTAL | ...
    - Filas 8+: datos
    - Filas finales: COSTO DIRECTO (SIN IGV), SUB TOTAL, IGV, TOTAL (CON IGV)

    Extrae:
    - expediente: TOTAL (CON IGV) de la seccion EXPEDIENTE
    - ppto_meta_hg: TOTAL (CON IGV) de la seccion PPTO META HG
    - monto_cc: TOTAL (CON IGV) de un PROVEEDOR (columnas entre EXPEDIENTE y PPTO META HG)
    """
    resultado = {"monto_cc": "No especificado", "ppto_meta_hg": "No especificado", "expediente": "No especificado"}

    try:
        max_row = min(ws.max_row or 200, 200)
        max_col = min(ws.max_column or 50, 50)

        # ================================================================
        # PASO 1: Encontrar headers de seccion (PPTO META HG y EXPEDIENTE)
        # ================================================================
        ppto_col_start = None
        ppto_row = None
        exp_col_start = None
        exp_row = None

        for row_idx in range(1, min(20, max_row + 1)):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().lower()
                    if ("ppto" in val and "meta" in val) or "meta hg" in val:
                        ppto_col_start = col_idx
                        ppto_row = row_idx
                    elif val == "expediente" or "expediente" in val:
                        exp_col_start = col_idx
                        exp_row = row_idx

        if not ppto_col_start:
            return resultado

        # ================================================================
        # PASO 2: Encontrar columnas "SUB TOTAL" bajo cada seccion
        # ================================================================
        header_row = ppto_row  # Fila de headers de seccion

        def _encontrar_subtotal_col(section_col_start, section_col_end):
            """Busca columna SUB TOTAL dentro de una seccion."""
            for r in range(header_row, min(header_row + 5, max_row + 1)):
                for c in range(section_col_start, min(section_col_end, max_col + 1)):
                    cell = ws.cell(row=r, column=c)
                    if cell.value and isinstance(cell.value, str):
                        v = cell.value.strip().lower()
                        if "sub total" in v or "subtotal" in v or v == "sub-total":
                            return c
            # Fallback: buscar columna numerica
            for c in range(section_col_start, min(section_col_end, max_col + 1)):
                for r in range(header_row + 1, min(header_row + 10, max_row + 1)):
                    val = ws.cell(row=r, column=c).value
                    if val and isinstance(val, (int, float)) and val >= 100:
                        return c
            return None

        ppto_subtotal_col = _encontrar_subtotal_col(ppto_col_start, ppto_col_start + 12)

        exp_subtotal_col = None
        if exp_col_start:
            # Limite de columnas de EXPEDIENTE: hasta la siguiente seccion o +12
            exp_end = ppto_col_start if ppto_col_start > exp_col_start else exp_col_start + 12
            exp_subtotal_col = _encontrar_subtotal_col(exp_col_start, exp_end)

        # ================================================================
        # PASO 3: Buscar TOTAL (CON IGV) para cada seccion
        # ================================================================
        def _extraer_total_igv(subtotal_col, section_col_start, section_end):
            """Busca el valor TOTAL (CON IGV) en la columna SUB TOTAL de una seccion."""
            if not subtotal_col:
                return None
            # Buscar desde abajo hacia arriba
            for row_idx in range(max_row, header_row, -1):
                fila_tiene_total_igv = False
                for check_col in range(1, max_col + 1):
                    cell_val = ws.cell(row=row_idx, column=check_col).value
                    if cell_val and isinstance(cell_val, str):
                        txt = cell_val.strip().lower()
                        if "total" in txt and "igv" in txt:
                            fila_tiene_total_igv = True
                            break

                if fila_tiene_total_igv:
                    val = ws.cell(row=row_idx, column=subtotal_col).value
                    if val and isinstance(val, (int, float)) and val >= 100:
                        return val
                    # Buscar en columnas cercanas de la seccion
                    for try_col in range(section_col_start, min(section_end, max_col + 1)):
                        val = ws.cell(row=row_idx, column=try_col).value
                        if val and isinstance(val, (int, float)) and val >= 100:
                            return val
                    break

            # Fallback: buscar SUB TOTAL o COSTO DIRECTO
            for row_idx in range(max_row, header_row, -1):
                for check_col in range(1, max_col + 1):
                    cell_val = ws.cell(row=row_idx, column=check_col).value
                    if cell_val and isinstance(cell_val, str):
                        txt = cell_val.strip().lower()
                        if txt in ["sub total", "subtotal", "costo directo", "costo directo (sin igv)"]:
                            val = ws.cell(row=row_idx, column=subtotal_col).value
                            if val and isinstance(val, (int, float)) and val >= 100:
                                return val
            # Ultimo fallback: ultimo valor grande
            last_val = None
            for row_idx in range(header_row + 2, max_row + 1):
                val = ws.cell(row=row_idx, column=subtotal_col).value
                if val and isinstance(val, (int, float)) and val >= 100:
                    last_val = val
            return last_val

        # PPTO META HG
        ppto_val = _extraer_total_igv(ppto_subtotal_col, ppto_col_start, ppto_col_start + 12)
        if ppto_val:
            resultado["ppto_meta_hg"] = f"S/ {ppto_val:,.2f}"

        # EXPEDIENTE
        if exp_subtotal_col:
            exp_end = ppto_col_start if ppto_col_start > exp_col_start else exp_col_start + 12
            exp_val = _extraer_total_igv(exp_subtotal_col, exp_col_start, exp_end)
            if exp_val:
                resultado["expediente"] = f"S/ {exp_val:,.2f}"

        # ================================================================
        # PASO 4: Buscar Monto CC (ganador/proveedor)
        # Columnas ENTRE EXPEDIENTE y PPTO META HG (donde estan los proveedores)
        # ================================================================
        if resultado["monto_cc"] == "No especificado":
            # Determinar rango de columnas de proveedores (excluir EXPEDIENTE y PPTO META HG)
            if exp_col_start and exp_subtotal_col:
                # Proveedores empiezan despues de la ultima columna de EXPEDIENTE
                prov_col_start = exp_subtotal_col + 1
            elif exp_col_start:
                prov_col_start = exp_col_start + 5  # estimacion
            else:
                prov_col_start = 1

            prov_col_end = ppto_col_start  # Proveedores terminan antes de PPTO META HG

            if prov_col_start < prov_col_end:
                for row_idx in range(max_row, 0, -1):
                    fila_tiene_total_igv = False
                    for check_col in range(1, max_col + 1):
                        cell_val = ws.cell(row=row_idx, column=check_col).value
                        if cell_val and isinstance(cell_val, str):
                            txt = cell_val.strip().lower()
                            if "total" in txt and "igv" in txt:
                                fila_tiene_total_igv = True
                                break

                    if fila_tiene_total_igv:
                        # Buscar valor en columnas de PROVEEDORES (entre EXPEDIENTE y PPTO)
                        for col_idx in range(prov_col_start, prov_col_end):
                            val = ws.cell(row=row_idx, column=col_idx).value
                            if val and isinstance(val, (int, float)) and val >= 100:
                                resultado["monto_cc"] = f"S/ {val:,.2f}"
                                break
                        break

    except Exception as e:
        print(f"    [WARN] Error leyendo hoja VS: {e}")

    return resultado


# ============================================================================
# EXTRACCION DE LINKS DE DRIVE
# ============================================================================

def _extraer_drive_links(texto):
    """Extrae links de Google Drive del cuerpo del correo (texto + HTML).
    Maneja URLs con /u/0/ o /u/N/ (cuenta de usuario de Google Workspace).
    """
    patrones = [
        # Carpetas - con o sin /u/N/
        r"https?://drive\.google\.com/drive(?:/u/\d+)?/folders/([a-zA-Z0-9_-]+)",
        # Archivos - con o sin /u/N/
        r"https?://drive\.google\.com(?:/u/\d+)?/file/d/([a-zA-Z0-9_-]+)",
        # Open por ID
        r"https?://drive\.google\.com(?:/u/\d+)?/open\?id=([a-zA-Z0-9_-]+)",
        # Google Sheets
        r"https?://docs\.google\.com(?:/u/\d+)?/spreadsheets/d/([a-zA-Z0-9_-]+)",
    ]

    links = []
    seen_ids = set()

    for patron in patrones:
        matches = re.findall(patron, texto)
        for match_id in matches:
            if match_id in seen_ids:
                continue
            seen_ids.add(match_id)

            if "folders" in patron:
                links.append({"type": "folder", "id": match_id})
            elif "spreadsheets" in patron:
                links.append({"type": "sheet", "id": match_id})
            else:
                links.append({"type": "file", "id": match_id})

    return links


# ============================================================================
# LECTURA DESDE GOOGLE DRIVE
# ============================================================================

def _leer_desde_drive(drive_service, sheets_service, link_info, asunto=""):
    """Lee datos de un archivo o carpeta de Drive."""
    resultado = {"monto_cc": "No especificado", "ppto_meta_hg": "No especificado", "expediente": "No especificado"}

    try:
        if link_info["type"] == "folder":
            resultado = _leer_carpeta_drive(drive_service, sheets_service, link_info["id"], asunto=asunto)
        elif link_info["type"] == "sheet":
            resultado = _leer_google_sheet(sheets_service, link_info["id"])
        elif link_info["type"] == "file":
            resultado = _leer_archivo_drive(drive_service, sheets_service, link_info["id"])
    except Exception as e:
        print(f"    [WARN] Error leyendo Drive ({link_info['type']}): {e}")

    return resultado


def _extraer_palabras_clave_asunto(asunto):
    """
    Extrae palabras clave del asunto del correo para matchear con nombres de archivo.
    Elimina prefijos como Re:, Fwd:, palabras genericas, y retorna palabras significativas.
    """
    if not asunto:
        return []

    texto = asunto.lower()
    # Quitar prefijos de respuesta (Re:, Fwd:, etc.)
    texto = re.sub(r"^(re|fwd|rv|fw)\s*:\s*", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"^(re|fwd|rv|fw)\s*:\s*", "", texto, flags=re.IGNORECASE)  # doble por "Re: Re:"

    # Quitar palabras genericas que no ayudan a matchear
    genericas = {
        "requerimiento", "cc", "envio", "adjunto", "solicitud", "cotizacion",
        "cotización", "propuesta", "para", "aprobacion", "aprobación", "de",
        "del", "por", "con", "los", "las", "una", "uno", "que", "se", "en",
        "la", "el", "al", "y", "o", "a", "no", "si", "su", "re", "fwd",
        "comparativo", "comparativos", "cuadro", "evaluacion", "evaluación",
        # Codigos de proyecto (no discriminan entre comparativos)
        "beethoven", "btv", "mater", "mara", "roosevelt", "alma",
    }

    # Extraer palabras de al menos 3 caracteres
    palabras = re.findall(r"[a-záéíóúñü]{3,}", texto)
    significativas = [p for p in palabras if p not in genericas]

    return significativas


def _calcular_score_match(nombre_archivo, palabras_clave):
    """
    Calcula un score de coincidencia entre un nombre de archivo y las palabras clave del asunto.
    Mayor score = mejor match.
    """
    if not palabras_clave:
        return 0

    nombre_lower = nombre_archivo.lower()
    score = 0

    for palabra in palabras_clave:
        if palabra in nombre_lower:
            score += 1
            # Bonus si la palabra es larga (mas especifica)
            if len(palabra) >= 6:
                score += 1

    return score


def _leer_carpeta_drive(drive_service, sheets_service, folder_id, asunto=""):
    """
    Lista archivos en una carpeta de Drive y busca el Excel del comparativo.
    Usa el asunto del correo para identificar el archivo correcto por nombre.
    """
    resultado = {"monto_cc": "No especificado", "ppto_meta_hg": "No especificado", "expediente": "No especificado"}

    try:
        response = drive_service.files().list(
            q=f"'{folder_id}' in parents and trashed = false",
            fields="files(id, name, mimeType)",
            pageSize=50,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()

        files = response.get("files", [])

        if not files:
            return resultado

        # Filtrar solo archivos de hojas de calculo
        spreadsheet_files = []
        for f in files:
            mime = f.get("mimeType", "")
            name_lower = f["name"].lower()
            es_hoja = (
                mime == "application/vnd.google-apps.spreadsheet"
                or any(mime.endswith(ext) for ext in ["spreadsheetml.sheet", "ms-excel", "openxmlformats"])
                or name_lower.endswith((".xlsx", ".xls", ".xlsm"))
            )
            if es_hoja:
                spreadsheet_files.append(f)

        if not spreadsheet_files:
            return resultado

        # Extraer palabras clave del asunto para matchear
        palabras_clave = _extraer_palabras_clave_asunto(asunto)

        # Calcular cuantos archivos contiene cada keyword (para pesar por discriminacion)
        keyword_frequency = {}
        for palabra in palabras_clave:
            count = sum(1 for f in spreadsheet_files if palabra in f["name"].lower())
            keyword_frequency[palabra] = max(count, 1)

        # Construir bigrams (pares de palabras consecutivas) para matching de frases
        bigrams = []
        for i in range(len(palabras_clave) - 1):
            bigrams.append(f"{palabras_clave[i]} {palabras_clave[i+1]}")
            # Tambien sin espacio (para nombres como "UPS-TRANSFORMADOR")
            bigrams.append(f"{palabras_clave[i]}-{palabras_clave[i+1]}")
            bigrams.append(f"{palabras_clave[i]}{palabras_clave[i+1]}")

        # Calcular score de match para cada archivo (keywords raras valen mas)
        archivos_con_score = []
        for f in spreadsheet_files:
            score = 0
            name_lower = f["name"].lower()
            # Score por keywords individuales
            for palabra in palabras_clave:
                if palabra in name_lower:
                    freq = keyword_frequency.get(palabra, 1)
                    peso = 3.0 / freq
                    score += peso
                    if len(palabra) >= 6:
                        score += 1.0 / freq
            # Bonus por bigrams (frases de 2 palabras consecutivas)
            for bigram in bigrams:
                if bigram in name_lower:
                    score += 5.0  # bonus fuerte por frase exacta
            # Bonus menor si tiene "comparativo" o "cuadro" en el nombre
            if "comparativo" in name_lower or "cuadro" in name_lower:
                score += 0.5
            # Tiebreaker: ratio de keywords matcheadas vs palabras en filename
            # Archivos con nombre mas corto/especifico ganan empates
            name_words = re.findall(r"[a-záéíóúñü]{3,}", name_lower)
            name_sig_words = [w for w in name_words if w not in {"rcos", "cuadro", "comparativo"}]
            matched_count = sum(1 for p in palabras_clave if p in name_lower)
            if name_sig_words and matched_count > 0:
                ratio = matched_count / len(name_sig_words)
                score += ratio * 2.0  # Bonus proporcional al ratio
            archivos_con_score.append((score, f))

        # Ordenar por score descendente (mejor match primero)
        archivos_con_score.sort(key=lambda x: x[0], reverse=True)

        # Log para debug
        if palabras_clave:
            top = archivos_con_score[0] if archivos_con_score else None
            if top and top[0] > 0:
                print(f"\n      [DRIVE] Carpeta con {len(files)} archivos, {len(spreadsheet_files)} hojas")
                print(f"      [DRIVE] Palabras clave: {palabras_clave[:5]}")
                print(f"      [DRIVE] Mejor match: '{top[1]['name']}' (score: {top[0]})")

        # Procesar archivos en orden de relevancia (max 3 para no tardar mucho)
        for score, f in archivos_con_score[:3]:
            mime = f.get("mimeType", "")
            datos = None

            try:
                if mime == "application/vnd.google-apps.spreadsheet":
                    datos = _leer_google_sheet(sheets_service, f["id"])
                else:
                    datos = _descargar_y_leer_excel(drive_service, f["id"], f["name"])
            except Exception as e:
                print(f"      [DRIVE] Error con '{f['name']}': {e}")
                continue

            if datos:
                if resultado["monto_cc"] == "No especificado" and datos.get("monto_cc") != "No especificado":
                    resultado["monto_cc"] = datos["monto_cc"]
                if resultado["ppto_meta_hg"] == "No especificado" and datos.get("ppto_meta_hg") != "No especificado":
                    resultado["ppto_meta_hg"] = datos["ppto_meta_hg"]
                if resultado["expediente"] == "No especificado" and datos.get("expediente") != "No especificado":
                    resultado["expediente"] = datos["expediente"]

            if resultado["monto_cc"] != "No especificado" and resultado["ppto_meta_hg"] != "No especificado":
                break

    except Exception as e:
        print(f"    [WARN] Error listando carpeta Drive: {e}")

    return resultado


# ============================================================================
# LECTURA DESDE GOOGLE SHEETS API
# ============================================================================

def _leer_google_sheet(sheets_service, spreadsheet_id):
    """Lee un Google Sheet buscando PPTO META HG y EXPEDIENTE en la pestaña VS."""
    resultado = {"monto_cc": "No especificado", "ppto_meta_hg": "No especificado", "expediente": "No especificado"}

    try:
        spreadsheet = sheets_service.spreadsheets().get(
            spreadsheetId=spreadsheet_id
        ).execute()

        sheets = spreadsheet.get("sheets", [])

        # Buscar pestaña "VS" primero (prioridad exacta)
        target_sheets = []
        for sheet in sheets:
            title = sheet["properties"]["title"]
            title_lower = title.lower().strip()
            if title_lower == "vs":
                target_sheets.insert(0, title)
            elif "vs" in title_lower:
                target_sheets.insert(1 if target_sheets else 0, title)
            elif any(kw in title_lower for kw in ["comparativo", "resumen", "cuadro"]):
                target_sheets.append(title)

        if not target_sheets:
            target_sheets = [s["properties"]["title"] for s in sheets[:3]]

        for sheet_title in target_sheets:
            try:
                # Leer rango amplio (hasta columna AZ fila 100)
                range_name = f"'{sheet_title}'!A1:AZ100"
                result = sheets_service.spreadsheets().values().get(
                    spreadsheetId=spreadsheet_id,
                    range=range_name,
                    valueRenderOption="UNFORMATTED_VALUE",
                ).execute()

                values = result.get("values", [])
                datos = _analizar_hoja_vs_sheets(values)

                if datos.get("ppto_meta_hg") != "No especificado":
                    resultado["ppto_meta_hg"] = datos["ppto_meta_hg"]
                if datos.get("monto_cc") != "No especificado":
                    resultado["monto_cc"] = datos["monto_cc"]
                if datos.get("expediente") != "No especificado":
                    resultado["expediente"] = datos["expediente"]

                if resultado["ppto_meta_hg"] != "No especificado":
                    break

            except Exception as e:
                continue

    except Exception as e:
        print(f"    [WARN] Error leyendo Google Sheet: {e}")

    return resultado


def _analizar_hoja_vs_sheets(values):
    """
    Analiza datos de la hoja VS (Google Sheets API - lista de listas).
    Misma logica que _leer_hoja_vs pero para datos del Sheets API.
    """
    resultado = {"monto_cc": "No especificado", "ppto_meta_hg": "No especificado", "expediente": "No especificado"}

    if not values:
        return resultado

    num_rows = len(values)
    max_cols = max(len(r) for r in values) if values else 0

    # ================================================================
    # PASO 1: Encontrar headers de seccion (PPTO META HG y EXPEDIENTE)
    # ================================================================
    ppto_col_start = None
    ppto_row = None
    exp_col_start = None
    exp_row = None

    for row_idx in range(min(20, num_rows)):
        row = values[row_idx]
        for col_idx, cell in enumerate(row):
            if cell and isinstance(cell, str):
                val = cell.strip().lower()
                if ("ppto" in val and "meta" in val) or "meta hg" in val:
                    ppto_col_start = col_idx
                    ppto_row = row_idx
                elif val == "expediente" or "expediente" in val:
                    exp_col_start = col_idx
                    exp_row = row_idx

    if ppto_col_start is None:
        return resultado

    header_row = ppto_row

    # ================================================================
    # PASO 2: Encontrar columnas "SUB TOTAL" bajo cada seccion
    # ================================================================
    def _encontrar_subtotal_col_sheets(section_start, section_end):
        for r_idx in range(header_row, min(header_row + 5, num_rows)):
            row = values[r_idx]
            for c_idx in range(section_start, min(section_end, len(row))):
                cell = row[c_idx]
                if cell and isinstance(cell, str):
                    v = cell.strip().lower()
                    if "sub total" in v or "subtotal" in v or v == "sub-total":
                        return c_idx
        # Fallback numerico
        for c_idx in range(section_start, min(section_end, max_cols)):
            for r_idx in range(header_row + 1, min(header_row + 10, num_rows)):
                row = values[r_idx]
                if c_idx < len(row):
                    num = _parsear_numero(row[c_idx])
                    if num and num >= 100:
                        return c_idx
        return None

    ppto_subtotal_col = _encontrar_subtotal_col_sheets(ppto_col_start, ppto_col_start + 12)

    exp_subtotal_col = None
    if exp_col_start is not None:
        exp_end = ppto_col_start if ppto_col_start > exp_col_start else exp_col_start + 12
        exp_subtotal_col = _encontrar_subtotal_col_sheets(exp_col_start, exp_end)

    # ================================================================
    # PASO 3: Buscar TOTAL (CON IGV) para cada seccion
    # ================================================================
    def _extraer_total_igv_sheets(subtotal_col, section_start, section_end):
        if subtotal_col is None:
            return None
        for row_idx in range(num_rows - 1, header_row, -1):
            row = values[row_idx]
            fila_tiene_total_igv = False
            for cell in row:
                if cell and isinstance(cell, str):
                    txt = cell.strip().lower()
                    if "total" in txt and "igv" in txt:
                        fila_tiene_total_igv = True
                        break
            if fila_tiene_total_igv:
                if subtotal_col < len(row):
                    num = _parsear_numero(row[subtotal_col])
                    if num and num >= 100:
                        return num
                for try_col in range(section_start, min(section_end, len(row))):
                    num = _parsear_numero(row[try_col])
                    if num and num >= 100:
                        return num
                break
        # Fallback: buscar SUB TOTAL / COSTO DIRECTO
        for row_idx in range(num_rows - 1, header_row, -1):
            row = values[row_idx]
            for cell in row:
                if cell and isinstance(cell, str):
                    txt = cell.strip().lower()
                    if txt in ["sub total", "subtotal", "costo directo", "costo directo (sin igv)"]:
                        if subtotal_col < len(row):
                            num = _parsear_numero(row[subtotal_col])
                            if num and num >= 100:
                                return num
        # Ultimo fallback
        last_val = None
        for row_idx in range(header_row + 2, num_rows):
            row = values[row_idx]
            if subtotal_col < len(row):
                num = _parsear_numero(row[subtotal_col])
                if num and num >= 100:
                    last_val = num
        return last_val

    # PPTO META HG
    ppto_val = _extraer_total_igv_sheets(ppto_subtotal_col, ppto_col_start, ppto_col_start + 12)
    if ppto_val:
        resultado["ppto_meta_hg"] = f"S/ {ppto_val:,.2f}"

    # EXPEDIENTE
    if exp_subtotal_col is not None:
        exp_end = ppto_col_start if ppto_col_start > exp_col_start else exp_col_start + 12
        exp_val = _extraer_total_igv_sheets(exp_subtotal_col, exp_col_start, exp_end)
        if exp_val:
            resultado["expediente"] = f"S/ {exp_val:,.2f}"

    # ================================================================
    # PASO 4: Buscar Monto CC (proveedor - entre EXPEDIENTE y PPTO META HG)
    # ================================================================
    if resultado["monto_cc"] == "No especificado" and ppto_col_start is not None:
        if exp_col_start is not None and exp_subtotal_col is not None:
            prov_col_start = exp_subtotal_col + 1
        elif exp_col_start is not None:
            prov_col_start = exp_col_start + 5
        else:
            prov_col_start = 0

        prov_col_end = ppto_col_start

        if prov_col_start < prov_col_end:
            for row_idx in range(num_rows - 1, 0, -1):
                row = values[row_idx]
                fila_tiene_total_igv = False
                for cell in row:
                    if cell and isinstance(cell, str):
                        txt = cell.strip().lower()
                        if "total" in txt and "igv" in txt:
                            fila_tiene_total_igv = True
                            break
                if fila_tiene_total_igv:
                    for col_idx in range(prov_col_start, min(prov_col_end, len(row))):
                        num = _parsear_numero(row[col_idx])
                        if num and num >= 100:
                            resultado["monto_cc"] = f"S/ {num:,.2f}"
                            break
                    break

    return resultado


# ============================================================================
# LECTURA DE ARCHIVOS DE DRIVE
# ============================================================================

def _descargar_y_leer_excel(drive_service, file_id, filename):
    """Descarga un archivo Excel de Drive y lo procesa."""
    try:
        request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
        file_data = request.execute()
        return _procesar_excel(file_data, filename)
    except Exception as e:
        print(f"    [WARN] Error descargando Excel de Drive: {e}")
        return None


def _leer_archivo_drive(drive_service, sheets_service, file_id):
    """Lee un archivo individual de Drive."""
    try:
        file_meta = drive_service.files().get(
            fileId=file_id, fields="id,name,mimeType",
            supportsAllDrives=True
        ).execute()

        mime = file_meta.get("mimeType", "")
        if mime == "application/vnd.google-apps.spreadsheet":
            return _leer_google_sheet(sheets_service, file_id)
        elif "spreadsheet" in mime or "excel" in mime:
            return _descargar_y_leer_excel(drive_service, file_id, file_meta["name"])
    except Exception as e:
        print(f"    [WARN] Error leyendo archivo Drive: {e}")

    return {"monto_cc": "No especificado", "ppto_meta_hg": "No especificado", "expediente": "No especificado"}


# ============================================================================
# UTILIDADES
# ============================================================================

def _parsear_numero(valor):
    """
    Parsea un valor a numero, manejando formatos:
    - Español: 39.488,25 (punto=miles, coma=decimal)
    - Ingles: 39,488.25 (coma=miles, punto=decimal)
    - Sin separador: 39488.25
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        limpio = valor.strip()
        # Remover simbolo de moneda
        limpio = re.sub(r"[sS]/\.?\s*", "", limpio)
        limpio = re.sub(r"(PEN|USD|US\$|\$)\s*", "", limpio)
        # Remover cualquier caracter que no sea digito, punto, coma o guion
        limpio = re.sub(r"[^\d.,\-]", "", limpio)

        if not limpio or limpio in ["-", ".", ","]:
            return None

        # Detectar formato por posicion del ultimo separador
        last_dot = limpio.rfind(".")
        last_comma = limpio.rfind(",")

        if last_dot > last_comma:
            # Punto es decimal (formato ingles): 39,488.25
            limpio = limpio.replace(",", "")
        elif last_comma > last_dot:
            # Coma es decimal (formato español): 39.488,25
            limpio = limpio.replace(".", "").replace(",", ".")
        else:
            # Solo un tipo de separador o ninguno
            limpio = limpio.replace(",", "")

        try:
            return float(limpio)
        except ValueError:
            return None
    return None
